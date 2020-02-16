import openpyxl

wk = openpyxl.load_workbook('C://Users/HaiYen/Desktop/TestData.xlsx')


def fetch_number_of_rows(SheetName):
    sh = wk[SheetName]
    return sh.max_row


def fetch_cell_data(SheetName, row, cell):
    sh = wk[SheetName]
    cell = sh.cell(int(row), int(cell))
    return cell.value


