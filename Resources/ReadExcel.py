import openpyxl

vk = openpyxl.load_workbook('C:\\Users\HaiYen\Desktop\TestData.xlsx')

print(vk.sheetnames)

print("Active sheet is " + vk.active.title)

sh = vk['Sheet1']

print(sh.title)

print(sh['A3'].value)

print(sh.cell(3, 1).value)

rows = sh.max_row
columns = sh.max_column

print("Total row is - " + str(rows))
print("Total column is - " + str(columns))

for i in range(1, rows + 1):
    for j in range(1, columns + 1):
        c = sh.cell(i, j)
        print(c.value)

for r in sh['A1':'C4']:
    for c in r:
        print(c.value)

# Write data in excel

vk = openpyxl.Workbook()
sh = vk.active
sh.title = 'Hello Yen Testing'

sh['A4'].value = 'T love you'
vk.create_sheet(title='hello somebody')

sh1 = vk['hello somebody']
sh1['A3'].value ='Who'

vk.save('C:\\Users\HaiYen\Desktop\WriteTestData.xlsx')
